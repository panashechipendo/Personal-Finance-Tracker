# Gab relevant packages
from tabulate import tabulate
from openpyxl import Workbook
import csv
import sys
import datetime


options = ["Add Transaction", "View Balance", "Set Budget", "View Budget Status", "Generate Report", "Export Data", "Reset", "Exit"]

# greeting screen
def display_menu():
    print("\n" + "="*30)
    print("       Finance Tracker")
    print("="*30)
    for i in range(len(options)):
        print(f"{i+1}. {options[i]}")
    print("="*30)

# Present options
def main():
    while True:
        display_menu()
        try:
            user_input = input("Choose a number from the options: ")
            choice = int(user_input)

            if choice < 1 or choice > len(options):
                print("Please enter a number between 1 and 8")
                continue
                
            if choice == 1:
                add_transaction()
            elif choice == 2:
                view_balance()
            elif choice == 3:
                set_budget()
            elif choice == 4: 
                view_budget_status()
            elif choice == 5:
                generate_report()
            elif choice == 6:
                export_data()
            elif choice == 7:
                reset()
            elif choice == 8:
                print("Thank you for using Finance Tracker!")
                sys.exit(0)
        except ValueError:
            print("Not a valid input")
        except KeyboardInterrupt:
            print("\nGoodbye!")
            sys.exit(0)
        except EOFError:
            print("\nGoodbye!")
            sys.exit(0)


# record transaction
def add_transaction():
    data = load_data()
    budget_data = load_budget_data()
    if len(data) == 0:
        balance = 0
    elif len(data) > 0:
        balance = data[-1]["balance"]

    while True:
        try:
            today = datetime.date
            transaction = input("Enter transaction types (income/expense): ").lower()
            if transaction == "income" or transaction == "expense":
                category = input("Enter category: ").upper()
            else:
                 continue
            
            if category == "":
                print("Please input a category(food, bills, etc...)")
        
            description = input("Enter description: ")

            if description == "":
                print("Please input a description preferably something describing the purchase")
                continue

            amount = input("Enter amount: ")
            if int(amount) <= 0:
                print("cant pay zero or a negative number for transaction")
                continue


            if transaction == "income":
                balance = int(balance) + int(amount)
            elif transaction == "expense":
                check_budget_alerts(category, amount)
                balance = int(balance) - int(amount)
            else:
                balance = balance
          
            savedata(today.today(), transaction, category, description, amount, balance)
            print("✔️ Transaction added successfully!")
            print()
            view_balance()
            break
        except KeyboardInterrupt:
            print("Thank you for using Finance Tracker!")
        except ValueError:
            print("Thats an invalid input")
            continue
            

def view_balance():
    data = load_data()
    if len(data) > 0:
        balance = data[-1]["balance"]
    else: 
        balance = 0

    print(f"Current balance: {balance}")
    continue_prompt = input("Press Enter to continue!: ")
    

# display budget
def view_budget_status():
    budget_remaining = calculate_budget_remaining()
    if len(budget_remaining) == 0:
        print("Cant display budget that's not defined")
        return
    
    print(tabulate(budget_remaining, headers="keys", tablefmt="grid"))
    continue_prompt = input("Press Enter to continue!: ")
    

# display transaction history
def generate_report():
    data = load_data()
    if len(data) > 0:
        print(tabulate(data, headers="keys", tablefmt="grid"))
        continue_prompt = input("Press Enter to continue!: ")
        
    else:
        print("No Transaction History")

# Export files in one of two formats
def export_data():
    data = load_data()
    if len(data) == 0:
        print("Can't export an empty file!")
        return
    formatted  = input("Please select a format(.xlsx(excel), .csv(comma separated value))")
    if formatted == ".xlsx" or formatted == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws["A1"] = "Date"
        ws["B1"] = "types"
        ws["C1"] = "Category"
        ws["D1"] = "Description"
        ws["E1"] = "Amount"
        ws["F1"] = "Balance"

        for entry in data:
            ws.append(
                [
                    entry["date"],
                    entry["types"],
                    entry["category"],
                    entry["description"],
                    entry["amount"],
                    entry["balance"]
                ]
            )
        wb.save("output.xlsx")
        print("File Exported!: output.xlsx")
    elif formatted == ".csv":
        with open("output.csv", "w") as file:
            writer = csv.DictWriter(file, fieldnames=["date", "types", "category", "description", "amount", "balance"])
            writer.writeheader()
            for entry in data:
                writer.writerow(entry)
        print("File exported! output.csv")
    continue_prompt = input("Press Enter to continue!: ")
    


# generate array of transactions
def load_data():
        data = []
        try:
            with open("transactions.csv", "r") as file:
                reader = csv.DictReader(file)

                for row in reader:
                    data.append(
                        {
                            "date": row["date"],
                            "types": row["types"],
                            "category": row["category"],
                            "description": row["description"],
                            "amount": row["amount"],
                            "balance": row["balance"]
                        }
                    )
        except FileNotFoundError:
            print("file not found creating file")
            with open("transactions.csv", "w", newline="") as file:
                writer = csv.DictWriter(file, fieldnames=["date", "types", "category", "description", "amount", "balance"])
                writer.writeheader()
        return data

# Same as above but for budget
def load_budget_data():
    budgets = []
    try:
        with open("budgets.csv", "r") as file:
            reader = csv.DictReader(file)
            for row in reader:
                budgets.append({
                    "category": row["category"],
                    "monthly_budget": int(row["monthly_budget"])
                })
    except FileNotFoundError:
        print("No file in directory creating file now...")
        with open("budgets.csv", "w", newline='') as file:
                writer = csv.DictWriter(file, fieldnames=["category", "monthly_budget"])
                writer.writeheader() 
    return budgets

# Save to csv
def savedata(date, types, category, description, amount, balance):
    with open("transactions.csv", "a") as file:
        writer = csv.DictWriter(file, fieldnames=["date", "types", "category", "description", "amount", "balance"])
        writer.writerow(
            {
                "date": date,
                "types": types,
                "category": category,
                "description": description,
                "amount": amount,
                "balance": balance
            }
        )

def set_budget():
    while True:
        try:
            category = input("Enter budget category: ").upper()
            monthly = int(input("Enter monthly budget: "))
            with open("budgets.csv", "a") as file:
                writer = csv.DictWriter(file, fieldnames=["category", "monthly_budget"])
                writer.writerow(
                    {
                        "category": category,
                        "monthly_budget": monthly,
                    }
                )
            break
        except ValueError:
            print("invalid input")
            continue
        except KeyboardInterrupt:
            print("Thank you for using finance tracker!")

# Determine remaining budget
def calculate_budget_remaining():
    budgets = load_budget_data()
    transactions = load_data()

    results = []

    for budget in budgets:
        category = budget["category"]
        monthly_limit = budget["monthly_budget"]

        category_expenses = [t for t in transactions if t["types"] == "expense" and t["category"] == category]

        total_spent = sum(int(expense["amount"]) for expense in category_expenses)

        remaining = monthly_limit - total_spent

        results.append({
            "category": category,
            "monthly_budget": monthly_limit,
            "current_spent": total_spent,
            "remaining": remaining
        })

    return results

# Alert the user if reaching or reached budget limit
def check_budget_alerts(current_category, current_amount):
    budget_remain = calculate_budget_remaining()
    if budget_remain:
        for remain in budget_remain:
            total = remain["monthly_budget"]
            current_spent = remain["current_spent"]

            if current_category == remain["category"]:
                current_spent += int(current_amount)

            remaining_after_transaction = total - current_spent

            if remaining_after_transaction <= 0:
                print(f"Alert!: {remain['category']}'s budget has been reached/exceeded: ${remaining_after_transaction}")
            elif remaining_after_transaction <= (0.1 * total) and remaining_after_transaction > 0:
                print(f"WARNING: {remain['category']}'s budget is about to be reached: ${remaining_after_transaction}")
            
# Clear all records
def reset():
    try:
        confirm = input("Are you sure you want to reset all data? This will delete all transactions and budgets. (y/n): ")

        if confirm.lower() == "y" or confirm.lower() == "yes":
            # Clear transactions file
            with open("transactions.csv", "w", newline="") as file:
                writer = csv.DictWriter(file, fieldnames=["date", "types", "category", "description", "amount", "balance"])
                writer.writeheader()

            with open("budgets.csv", "w", newline='') as file:
                writer = csv.DictWriter(file, fieldnames=["category", "monthly_budget"])
                writer.writeheader() 

            print("✔️ All data has been reset successfully!")
        else:
            print("Reset cancelled.")
    except FileNotFoundError:
        print("No data files found to reset")
    except Exception as e:
        print(f"An error occurred while resetting: {e}")



if __name__ == "__main__":
    main()